from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import time
import datetime


executable_path= "C:/Users/Nianyi/PycharmProjects/webdriver/chromedriver.exe"
excel_path="C:/Users/Nianyi/PycharmProjects/pythonProject/house.xlsx"
property_search_website="https://hillsborough.county-taxes.com/public/search/property_tax?action_location=Landing+Page"

workbook = load_workbook(excel_path)
sheet = workbook.active
total_houses = sheet.max_row - 1

driver = webdriver.Chrome(executable_path)
driver.get(property_search_website)

for num in range(total_houses):
    #get address
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    address_row = num + 2
    house_address = sheet.cell(row=address_row, column=1).value

    #search address
    search_box = driver.find_element(By.ID, "search_query")
    search_box.send_keys(house_address)
    search_button = driver.find_element(By.XPATH,
                                        "/html/body/div[2]/div[2]/div/main/section/div[1]/form/div/div/div[2]/button").click()
    time.sleep(3)
    property_page = driver.current_window_handle
    title_name = driver.find_element(By.XPATH, "//*[@class=' container result property_tax']/div[1]/div/a").click()
    appr_link = driver.find_element(By.XPATH,
                                    "/html/body/div[2]/div[2]/div/main/section/div[2]/div[2]/div[3]/div[3]/a").click()
    time.sleep(3)
    info_page = driver.switch_to.window(driver.window_handles[1])
    time.sleep(1)
    beds = driver.find_element(By.XPATH, "//td[text()='Bedrooms']/following-sibling::td[1]").text
    sheet.cell(row=address_row, column=3).value = beds
    baths = driver.find_element(By.XPATH, "//td[text()='Bathrooms']/following-sibling::td[1]").text
    sheet.cell(row=address_row, column=4).value = baths
    stories = driver.find_element(By.XPATH, "//td[text()='Stories']/following-sibling::td[1]").text
    sheet.cell(row=address_row, column=5).value = stories
    area = driver.find_element(By.XPATH, "//th[contains(@data-bind, 'text: publicHeatedArea')]").text.replace(",", "")
    round_sqft = round(int(area), -2)
    sheet.cell(row=address_row, column=6).value = area
    price = driver.find_element(By.XPATH, "//td[contains(@data-bind, 'text: publicPrice')]").text.replace("$", "")
    sheet.cell(row=address_row, column=7).value = price

    if sheet.cell(row=address_row, column=2).value == None:
        month = int(driver.find_element(By.XPATH, "//td[contains(@data-bind, 'text: publicMonth')]").text)
        year = int(driver.find_element(By.XPATH, "//td[contains(@data-bind, 'text: publicYear')]").text)
        date = datetime.datetime(int(round(float(year))), int(round(float(month))), 15)
        sheet.cell(row=address_row, column=2).value = date.strftime("%x")

    driver.close()

    if beds == "4.0" and baths == "3.0":
        if 2900 < round_sqft < 3300:
            sheet.cell(row=address_row, column=8).value = "Avila"
            sheet.cell(row=address_row, column=9).value = "690990"
        if 2300 < round_sqft < 2700:
            sheet.cell(row=address_row, column=8).value = "Petaluma"
            sheet.cell(row=address_row, column=9).value = "610990"
    elif beds == "4.0" and baths == "3.5":
        if 3100 < round_sqft < 3350:
            sheet.cell(row=address_row, column=8).value = "Modesto"
            sheet.cell(row=address_row, column=9).value = "689990"
        elif 3350 <= round_sqft < 3700:
            sheet.cell(row=address_row, column=8).value = "Amberly"
            sheet.cell(row=address_row, column=9).value = "733490"
        elif 4000 <= round_sqft < 4400:
            sheet.cell(row=address_row, column=8).value = "Solana"
            sheet.cell(row=address_row, column=9).value = "767990"
        else:
            sheet.cell(row=address_row, column=8).value = "Sonama"
            sheet.cell(row=address_row, column=9).value = "711990"
    elif beds == "5.0" and baths == "4.5":
        sheet.cell(row=address_row, column=8).value = "Ventura"
        sheet.cell(row=address_row, column=9).value = "964990"
    elif beds == "5.0" and baths == "4.0":
        if 4100 < round_sqft < 4200:
            sheet.cell(row=address_row, column=8).value = "Bellejo"
            sheet.cell(row=address_row, column=9).value = "766490"
        elif 4200 <= round_sqft < 4300:
            sheet.cell(row=address_row, column=8).value = "Sonara"
            sheet.cell(row=address_row, column=9).value = "807990"
    elif beds == "5.0" and baths == "5.0":
        sheet.cell(row=address_row, column=8).value = "Mendocino"
        sheet.cell(row=address_row, column=9).value = "853990"
    elif beds == "6.0" and baths == "4.0":
        sheet.cell(row=address_row, column=8).value = "Seabrook"
        sheet.cell(row=address_row, column=9).value = "822990"
    elif beds == "5.0" and baths == "3.5":
        sheet.cell(row=address_row, column=8).value = "Daphne"
        sheet.cell(row=address_row, column=9).value = "819990"
    else:
        sheet.cell(row=address_row, column=8).value = "Unknown"

    print("wrote" + beds, baths, price, area + " to the excel sheet")
    workbook.save(excel_path)
    driver.switch_to.window(property_page)
    time.sleep(3)
    driver.find_element(By.XPATH, "/html/body/div[2]/header/div/section/div[2]/div/div/div[2]/div/a/h2/div[2]").click()
    time.sleep(2)
    print("Address has been added: " + house_address)

driver.close()


